import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import xml.etree.ElementTree as ET
import re
import os
import tempfile
import shutil
import datetime
import webbrowser
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from PIL import Image, ImageDraw, ImageFont, ImageTk


def is_ipv4(addr):
    return bool(re.fullmatch(r"\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}", addr))


def parse_l5x(filepath, progress_callback):
    tree = ET.parse(filepath)
    root = tree.getroot()

    modules = root.findall(".//Module")
    total = len(modules)

    results = []
    for i, module in enumerate(modules, 1):
        name = module.get("Name", "")
        catalog = module.get("CatalogNumber", "")

        ports = module.findall("Ports/Port")
        for port in ports:
            addr = port.get("Address", "")
            if is_ipv4(addr):
                results.append((name, catalog, addr))

        progress_callback(i, total)

    return results


class App:
    def __init__(self):
        self.window = tk.Tk()
        self.window.title("L5X Extractor - Módulos com IP")
        self.window.geometry("900x600")
        self.window.minsize(600, 300)

        self.filepath = tk.StringVar()
        self.spreadsheet_path = tk.StringVar()
        self.results = []
        self._pending_changes = []

        self._set_icon()
        self._build_menu()
        self._build_ui()

    def _set_icon(self):
        size = 64
        img = Image.new("RGBA", (size, size), (68, 114, 196, 255))
        draw = ImageDraw.Draw(img)
        try:
            font = ImageFont.truetype("arial.ttf", 28)
        except (IOError, OSError):
            font = ImageFont.load_default()
        text = "L5K"
        bbox = draw.textbbox((0, 0), text, font=font)
        x = (size - (bbox[2] - bbox[0])) / 2 - bbox[0]
        y = (size - (bbox[3] - bbox[1])) / 2 - bbox[1]
        draw.text((x, y), text, fill="white", font=font)
        tmp = tempfile.NamedTemporaryFile(suffix=".png", delete=False)
        img.save(tmp, format="PNG")
        tmp.close()
        self._icon = tk.PhotoImage(file=tmp.name)
        self.window.iconphoto(True, self._icon)
        os.unlink(tmp.name)

    def _build_menu(self):
        menu_bar = tk.Menu(self.window)
        self.window.config(menu=menu_bar)

        ajuda = tk.Menu(menu_bar, tearoff=0)
        menu_bar.add_cascade(label="Ajuda", menu=ajuda)
        ajuda.add_command(label="Ajuda", command=self._show_ajuda)
        ajuda.add_separator()
        ajuda.add_command(label="Sobre", command=self._show_sobre)

    def _show_ajuda(self):
        messagebox.showinfo(
            "Ajuda - UpdateIpList",
            "Como usar o UpdateIpList:\n\n"
            "1. Clique em 'Procurar L5X' e selecione um arquivo .L5X\n"
            "2. Aguarde o processamento (barra de progresso)\n"
            "3. Os módulos com endereço IP serão listados\n"
            "4. Clique em 'Procurar Planilha' e selecione a planilha existente\n"
            "5. Clique em 'Atualizar Planilha' para criar nova versão\n\n"
            "O programa extrai módulos com IP do L5X e atualiza\n"
            "a coluna Nomenclatura na planilha, destacando em\n"
            "azul (novo) ou amarelo (editado)."
        )

    def _show_sobre(self):
        win = tk.Toplevel(self.window)
        win.title("Sobre - UpdateIpList")
        win.geometry("380x220")
        win.resizable(False, False)
        win.transient(self.window)
        win.grab_set()

        frame = ttk.Frame(win, padding="20")
        frame.pack(fill=tk.BOTH, expand=True)

        ttk.Label(frame, text="UpdateIpList", font=("Segoe UI", 14, "bold")).pack()
        ttk.Label(frame, text="Versão 1.0").pack(pady=(5, 0))
        ttk.Label(frame, text="Autolinx Automação").pack(pady=(5, 0))
        ttk.Label(frame, text="Contato:").pack(pady=(10, 0))
        email_link = ttk.Label(frame, text="lewicks@gmail.com", foreground="blue", cursor="hand2")
        email_link.pack()
        email_link.bind("<Button-1>", lambda e: webbrowser.open("mailto:lewicks@gmail.com"))

        ttk.Button(frame, text="Fechar", command=win.destroy).pack(pady=(15, 0))

    def _build_ui(self):
        # --- L5X file selection ---
        frame_l5x = ttk.Frame(self.window, padding="5")
        frame_l5x.pack(fill=tk.X)

        ttk.Label(frame_l5x, text="Arquivo L5X:").pack(side=tk.LEFT)
        entry_l5x = ttk.Entry(frame_l5x, textvariable=self.filepath)
        entry_l5x.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(5, 5))
        ttk.Button(frame_l5x, text="Procurar L5X", command=self.browse_file).pack(side=tk.RIGHT)

        # --- spreadsheet selection ---
        frame_sheet = ttk.Frame(self.window, padding="5")
        frame_sheet.pack(fill=tk.X)

        ttk.Label(frame_sheet, text="Planilha:").pack(side=tk.LEFT)
        entry_sheet = ttk.Entry(frame_sheet, textvariable=self.spreadsheet_path)
        entry_sheet.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(5, 5))
        ttk.Button(frame_sheet, text="Procurar Planilha", command=self.browse_spreadsheet).pack(side=tk.RIGHT)

        # --- treeview ---
        frame_tree = ttk.Frame(self.window, padding="5")
        frame_tree.pack(fill=tk.BOTH, expand=True)

        columns = ("Name", "CatalogNumber", "Address")
        self.tree = ttk.Treeview(frame_tree, columns=columns, show="headings")

        self.tree.heading("Name", text="Name")
        self.tree.heading("CatalogNumber", text="CatalogNumber")
        self.tree.heading("Address", text="Address")

        self.tree.column("Name", width=300, minwidth=100)
        self.tree.column("CatalogNumber", width=250, minwidth=100)
        self.tree.column("Address", width=180, minwidth=100)

        scrollbar = ttk.Scrollbar(frame_tree, orient=tk.VERTICAL, command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar.set)

        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        # --- progress bar ---
        frame_progress = ttk.Frame(self.window, padding="5")
        frame_progress.pack(fill=tk.X)

        self.progress_var = tk.IntVar()
        self.progress_bar = ttk.Progressbar(frame_progress, variable=self.progress_var, maximum=100)
        self.progress_bar.pack(fill=tk.X, pady=(0, 2))

        self.progress_label = ttk.Label(frame_progress, text="")
        self.progress_label.pack(anchor=tk.W)

        # --- status bar ---
        frame_status = ttk.Frame(self.window, padding="5")
        frame_status.pack(fill=tk.X)

        self.status_label = ttk.Label(frame_status, text="Nenhum arquivo carregado.")
        self.status_label.pack(side=tk.LEFT)

        ttk.Button(frame_status, text="Atualizar Planilha", command=self.update_spreadsheet).pack(side=tk.RIGHT)

    def update_progress(self, current, total):
        if total > 0:
            pct = int(current / total * 100)
            self.progress_var.set(pct)
            self.progress_label.config(text=f"Processando módulo {current} de {total}... ({pct}%)")
            self.window.update_idletasks()

    def browse_file(self):
        path = filedialog.askopenfilename(
            title="Selecionar arquivo L5X",
            filetypes=[("L5X Files", "*.L5X"), ("All Files", "*.*")]
        )
        if not path:
            return

        self.filepath.set(path)
        self.tree.delete(*self.tree.get_children())
        self.results.clear()
        self.progress_var.set(0)
        self.progress_label.config(text="")
        self.status_label.config(text="Carregando...")
        self.window.update_idletasks()

        try:
            self.results = parse_l5x(path, self.update_progress)
        except ET.ParseError as e:
            messagebox.showerror("Erro de XML", f"Arquivo L5X inválido ou mal formatado:\n{e}")
            self.status_label.config(text="Erro: arquivo inválido.")
            self.progress_label.config(text="")
            return
        except FileNotFoundError:
            messagebox.showerror("Erro", "Arquivo não encontrado.")
            self.status_label.config(text="Erro: arquivo não encontrado.")
            self.progress_label.config(text="")
            return
        except Exception as e:
            messagebox.showerror("Erro", f"Falha ao processar arquivo:\n{e}")
            self.status_label.config(text="Erro ao processar arquivo.")
            self.progress_label.config(text="")
            return

        for row in self.results:
            self.tree.insert("", tk.END, values=row)

        count = len(self.results)
        self.status_label.config(text=f"Encontrados: {count} módulo(s) com IP")
        self.progress_label.config(text=f"Concluído! {count} módulo(s) com IP encontrado(s).")

    def browse_spreadsheet(self):
        path = filedialog.askopenfilename(
            title="Selecionar planilha existente",
            filetypes=[("Excel Files", "*.xlsx"), ("All Files", "*.*")]
        )
        if path:
            self.spreadsheet_path.set(path)

    def _reconstruct_ip(self, ws, row):
        h = ws.cell(row=row, column=8).value
        j = ws.cell(row=row, column=10).value
        l = ws.cell(row=row, column=12).value
        n = ws.cell(row=row, column=14).value
        try:
            return f"{int(h)}.{int(j)}.{int(l)}.{int(n)}"
        except (TypeError, ValueError):
            return None

    def update_spreadsheet(self):
        if not self.results:
            messagebox.showinfo("Aviso", "Nenhum dado do L5X para processar.")
            return

        orig_path = self.spreadsheet_path.get()
        if not orig_path:
            messagebox.showinfo("Aviso", "Selecione uma planilha existente primeiro.")
            return

        today = datetime.date.today()
        new_path = re.sub(
            r'_\d{4}_\d{2}_\d{2}\.xlsx$',
            f'_{today.year}_{today.month:02d}_{today.day:02d}.xlsx',
            orig_path
        )

        if new_path == orig_path:
            base, ext = os.path.splitext(orig_path)
            new_path = f"{base}_{today.year}_{today.month:02d}_{today.day:02d}{ext}"

        try:
            wb_data = load_workbook(orig_path, data_only=True)
        except Exception as e:
            messagebox.showerror("Erro", f"Não foi possível abrir a planilha:\n{e}")
            return

        conflicts = []
        updated = 0
        added = 0
        not_found = []

        sheets_to_process = [s for s in wb_data.sheetnames if s.startswith("PLC")]

        for name, catalog, ip in self.results:
            name_clean = name.lstrip("_")
            row_found = None
            sheet_found = None

            for sn in sheets_to_process:
                ws_data = wb_data[sn]
                for row_idx in range(12, ws_data.max_row + 1):
                    row_ip = self._reconstruct_ip(ws_data, row_idx)
                    if row_ip == ip:
                        row_found = row_idx
                        sheet_found = sn
                        break
                if row_found:
                    break

            if row_found is None:
                not_found.append((name, ip))
                continue

            current_val = wb_data[sheet_found].cell(row=row_found, column=5).value

            if current_val is None or (isinstance(current_val, str) and current_val.strip() == ""):
                self._pending_changes.append((sheet_found, row_found, name_clean, "blue"))
                added += 1
            elif current_val == name_clean:
                self._pending_changes.append((sheet_found, row_found, name_clean, "yellow"))
                updated += 1
            else:
                conflicts.append((ip, name_clean, str(current_val)))

        if not_found:
            msg = "IPs não encontrados na planilha:\n"
            for n, i in not_found:
                msg += f"  {n} ({i})\n"
            messagebox.showwarning("IPs não localizados", msg)

        if conflicts:
            msg = "Conflitos - dispositivos diferentes no mesmo IP:\n\n"
            for ip, l5x_name, sheet_name in conflicts:
                msg += f"IP {ip}:\n  L5X: {l5x_name}\n  Planilha: {sheet_name}\n\n"
            messagebox.showwarning("Conflitos encontrados", msg)

        if not self._pending_changes:
            messagebox.showinfo("Nenhuma alteração", "Nenhuma linha precisou ser alterada.")
            return

        try:
            shutil.copy2(orig_path, new_path)
        except Exception as e:
            messagebox.showerror("Erro", f"Falha ao copiar planilha:\n{e}")
            return

        try:
            wb = load_workbook(new_path)
        except Exception as e:
            messagebox.showerror("Erro", f"Falha ao abrir cópia:\n{e}")
            return

        blue_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")

        for sheet_name, row_idx, value, action in self._pending_changes:
            ws = wb[sheet_name]
            cell = ws.cell(row=row_idx, column=5)
            cell.value = value
            if action == "blue":
                cell.fill = blue_fill
            elif action == "yellow":
                cell.fill = yellow_fill

        try:
            wb.save(new_path)
        except Exception as e:
            messagebox.showerror("Erro", f"Falha ao salvar planilha:\n{e}")
            return

        self._pending_changes.clear()

        msg = f"Planilha atualizada com sucesso:\n{new_path}\n\n"
        msg += f"Novos (azul): {added}\n"
        msg += f"Editados (amarelo): {updated}\n"
        if not_found:
            msg += f"Não encontrados: {len(not_found)}\n"
        if conflicts:
            msg += f"Conflitos: {len(conflicts)}"
        messagebox.showinfo("Concluído", msg)

    def run(self):
        self.window.mainloop()


if __name__ == "__main__":
    App().run()
